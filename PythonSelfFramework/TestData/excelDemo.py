import openpyxl
import os
path = "D:\\gaurav\\SELENIUM with Python\\PythonSelfFramework\\TestData\\PythonDemo.xlsx"
book = openpyxl.load_workbook(path)
sheet = book.active
dict_1 = {}
cell = sheet.cell(row=1,column=2)
print(cell.value)

sheet.cell(row=2,column=2).value = "Rohit"
sheet.cell(row=3,column=2).value = "mohit"
print("for row 2 and column 2 value is: ",sheet.cell(row=2,column=2).value)
# another way to get the value of above line of code
print("for row 2 and column 2 value is: ",sheet['B2'].value)
print("for row 3 and column 2 value is: ",sheet.cell(row=3,column=2).value)
print("for row 3 and column 2 value is: ",sheet['B3'].value)
print("Maximum rows used in our excel sheet is: ",sheet.max_row)
print("Maximum columns used in our excel sheet is: ",sheet.max_column)
print(os.getcwd())
for r  in range(1,sheet.max_row+1):
    if sheet.cell(row=r,column=1).value == "Testcase2":
        for col in range(2,sheet.max_column +1):
            # print(sheet.cell(row=r,column=col).value)
            dict_1[sheet.cell(row=1,column=col).value] = sheet.cell(row=r,column=col).value

print(dict_1)


