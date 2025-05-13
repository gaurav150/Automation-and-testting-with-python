import time
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
file_path = r"C:\Users\Piyush\Downloads\download.xlsx"
fruit_name = "Apple"
new_value = "999"
def update_excel_data(file_path,searchTerm,colName,new_value):
    Dict = {}
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                Dict['row'] = i
    target = sheet.cell(row=Dict['row'], column=Dict['col']).value = new_value
    book.save(file_path)

driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID,"downloadButton").click()
## edit the Excel with updated value
update_excel_data(file_path,fruit_name,"price","999")

file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)
wait = WebDriverWait(driver,5)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)
price_column = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_column+"-undefined']").text
print(actual_price)
assert  actual_price == new_value
time.sleep(500)